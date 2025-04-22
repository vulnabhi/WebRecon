#!/usr/bin/env python3

import argparse
import requests
import re
import openpyxl # For Excel export
import os # For checking file existence
from urllib.parse import urljoin, urlparse
from bs4 import BeautifulSoup
# Import styles for Excel formatting
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter # For column width adjustment
# Imports for threading and summary count
import threading
import queue
from collections import Counter
# Import for running external commands and DNS resolution
import subprocess
import shlex
import socket # For DNS resolution
# Import for config file handling
import json

# --- Configuration ---

# Define the string to use when a technology category is not detected
NOT_DETECTED_STRING = "Not Found" # <--- CHANGE THIS STRING IF YOU WANT SOMETHING ELSE

# Flag to track if Shodan API error has been printed (to avoid repetition)
shodan_api_error_printed = False

# PATTERNS dictionaries remain unchanged (CMS, Backend, Frontend)
CMS_PATTERNS = {
    # Format: CMS Name: (Detected String with {version} placeholder, Version if found, Source Info)
    'WordPress': [
        # Meta tag is often the best source for version
        {'type': 'meta', 'name': 'generator', 'content': r'WordPress\s*([\d\.]+)?', 'version_group': 1},
        # Strong path/script indicators (check these early)
        {'type': 'path', 'pattern': '/wp-content/'}, # Check for core path
        {'type': 'path', 'pattern': '/wp-includes/'}, # Check for core path
        {'type': 'script', 'pattern': r'/wp-emoji-release\.min\.js'}, # Common WP script (Raw string fix)
        # Fallback version detection from query parameters
        {'type': 'link_script_ver', 'pattern': r'\?ver=([\d\.]+)'}, # Check ?ver= in script/link tags
        # Other indicators
        {'type': 'header', 'name': 'Link', 'pattern': 'rel="https://api.w.org/"'},
        {'type': 'html', 'pattern': 'wp-block-library-css'}, # Common CSS ID/handle
    ],
    'Joomla': [
        {'type': 'meta', 'name': 'generator', 'content': r'Joomla!\s*([\d\.]+)?', 'version_group': 1},
        {'type': 'path', 'pattern': '/administrator/'},
        {'type': 'cookie', 'name': re.compile(r'^[a-f0-9]{32}$')}, # Joomla session cookie pattern
        {'type': 'script', 'pattern': '/media/jui/js/jquery.min.js'}, # Often indicates Joomla, but check jQuery separately too
    ],
    'Drupal': [
        {'type': 'meta', 'name': 'generator', 'content': r'Drupal\s*([\d\.]+)?', 'version_group': 1},
        {'type': 'header', 'name': 'X-Generator', 'pattern': r'Drupal\s*([\d\.]+)?', 'version_group': 1},
        {'type': 'script', 'pattern': '/misc/drupal.js'},
        {'type': 'path', 'pattern': '/sites/default/'},
    ],
    'Shopify': [ # Version info not typically exposed easily
        {'type': 'header', 'name': 'X-Shopify-Stage', 'pattern': '.'},
        {'type': 'script', 'pattern': 'cdn.shopify.com'},
        {'type': 'html', 'pattern': 'Shopify.theme'},
    ],
    'Magento': [ # Version info not typically exposed easily
        {'type': 'path', 'pattern': '/media/wysiwyg/'},
        {'type': 'path', 'pattern': '/skin/frontend/'},
        {'type': 'script', 'pattern': 'mage/'},
        {'type': 'script', 'pattern': 'varien/'},
        {'type': 'cookie', 'name': 'frontend'},
    ]
    # Add more CMS patterns here
}

# Backend technology patterns with version capture groups
BACKEND_PATTERNS = {
    # Format: Tech Name: (Detected String with {version} placeholder, Version if found, Source Info)
    'PHP': [
        {'type': 'header', 'name': 'X-Powered-By', 'pattern': r'PHP/?([\d\.]+)?', 'version_group': 1}
    ],
    'ASP.NET': [
        {'type': 'header', 'name': 'X-Powered-By', 'pattern': r'ASP\.NET'},
        {'type': 'header', 'name': 'X-AspNet-Version', 'pattern': r'([\d\.]+)', 'version_group': 1}, # Captures version directly
        {'type': 'cookie', 'name': 'ASPSESSIONID'},
        {'type': 'cookie', 'name': 'ASP.NET_SessionId'},
    ],
    'Node.js': [
        {'type': 'header', 'name': 'X-Powered-By', 'pattern': r'Express'} # Express often reveals itself
    ],
    'Ruby on Rails': [
        {'type': 'header', 'name': 'X-Powered-By', 'pattern': r'(Phusion Passenger|mod_rails)(?:\s*\(?([\d\.]+)\)?)?', 'version_group': 2}
    ],
    'Django': [
        {'type': 'cookie', 'name': 'csrftoken'}
    ],
    'Java/JSP': [
        {'type': 'cookie', 'name': 'JSESSIONID'}
    ],
    'Nginx': [
        {'type': 'header', 'name': 'Server', 'pattern': r'nginx/?([\d\.]+)?', 'version_group': 1}
    ],
    'Apache': [
        # Order matters: Check for version in brackets first
        {'type': 'header', 'name': 'Server', 'pattern': r'Apache/([\d\.]+)\s+\(', 'version_group': 1},
        {'type': 'header', 'name': 'Server', 'pattern': r'Apache(?:/([\d\.]+))?', 'version_group': 1},
    ],
    'IIS': [
        {'type': 'header', 'name': 'Server', 'pattern': r'Microsoft-IIS/([\d\.]+)', 'version_group': 1}
    ],
    # Add more backend patterns here
}

# Frontend library patterns - Attempting version detection from filenames
# Note: This is less reliable than backend header checks.
FRONTEND_PATTERNS = {
    'React': [
        # Version often not in filename for React build tools
        {'type': 'script', 'pattern': r'react(\.min)?\.js|react-dom(\.min)?\.js'},
        {'type': 'html', 'pattern': 'data-reactroot|data-reactid'},
        {'type': 'script_content', 'pattern': r'React\.createElement'}
    ],
    'Angular': [
        # AngularJS (1.x) filename pattern
        {'type': 'script', 'pattern': r'angular(?:/|-)([\d\.]+)(?:/|-)?angular(\.min)?\.js', 'version_group': 1},
        {'type': 'script', 'pattern': r'angular(\.min)?\.js'}, # Generic fallback
        {'type': 'html', 'pattern': 'ng-app|ng-controller|ng-model'}, # More common in AngularJS (1.x)
        # Angular (2+) often uses build filenames like main.<hash>.js, harder to detect version from filename
        {'type': 'html', 'pattern': r'\sng-version="([\d\.]+)"', 'version_group': 1} # Check for ng-version attribute
    ],
    'Vue.js': [
        {'type': 'script', 'pattern': r'vue(?:/|-|@)([\d\.]+)(?:/|-)?vue(\.min)?\.js', 'version_group': 1},
        {'type': 'script', 'pattern': r'vue(\.min)?\.js'}, # Generic fallback
        {'type': 'html', 'pattern': 'v-app|v-bind|@click'},
        {'type': 'script_content', 'pattern': r'new Vue\('}
    ],
    'jQuery': [
        # Common filename patterns
        {'type': 'script', 'pattern': r'jquery(?:-|.)([\d\.]+(?:-\w+)?)(?:\.min)?\.js', 'version_group': 1}, # e.g., jquery-3.6.0.min.js, jquery.1.12.4.js
        {'type': 'script_content', 'pattern': r'jQuery\.fn\.jquery:\s*"([\d\.]+)"', 'version_group': 1}, # Check inline version variable
        {'type': 'script_content', 'pattern': r'jQuery|\$\('} # Generic fallback
    ],
    'Bootstrap': [
        # Common filename patterns for JS and CSS
        {'type': 'script', 'pattern': r'bootstrap(?:-|.)([\d\.]+)(?:-|.)?js', 'version_group': 1}, # e.g., bootstrap.5.3.0.bundle.min.js
        {'type': 'css', 'pattern': r'bootstrap(?:-|.)([\d\.]+)(?:-|.)?css', 'version_group': 1}, # e.g., bootstrap.min-5.3.css
        {'type': 'script', 'pattern': r'bootstrap(\.min)?\.js'}, # Generic fallback
        {'type': 'css', 'pattern': r'bootstrap(\.min)?\.css'}, # Generic fallback
        {'type': 'html', 'pattern': 'class="container|row|col-'}
    ],
    # Add more frontend patterns here
}

# --- Helper Functions ---

def make_request(url, verbose=False): # Added verbose flag
    """Makes an HTTP GET request to the URL."""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36 BackReconTool/1.9' # Version bump
    }
    try:
        requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)
    except AttributeError:
        pass

    session = requests.Session()
    session.headers.update(headers)
    session.max_redirects = 5

    try:
        parsed_url = urlparse(url)
        target_display = url[:50] + "..." if len(url) > 50 else url

        if not parsed_url.scheme:
            target_url_https = 'https://' + url
            if verbose: print(f"[*] [{target_display}] No scheme provided. Trying {target_url_https}")
            try:
                response = session.get(target_url_https, timeout=15, verify=True, allow_redirects=True)
                response.raise_for_status()
            except requests.exceptions.SSLError as ssl_err:
                 if verbose: print(f"[*] [{target_display}] HTTPS SSL error: {ssl_err}. Trying HTTP...")
                 target_url_http = f'http://{parsed_url.netloc}{parsed_url.path or "/"}' + (f"?{parsed_url.query}" if parsed_url.query else "")
                 response = session.get(target_url_http, timeout=15, verify=False, allow_redirects=True)
                 response.raise_for_status()
            except requests.exceptions.RequestException as req_err:
                if verbose: print(f"[*] [{target_display}] HTTPS failed: {req_err}. Trying HTTP...")
                target_url_http = f'http://{parsed_url.netloc}{parsed_url.path or "/"}' + (f"?{parsed_url.query}" if parsed_url.query else "")
                response = session.get(target_url_http, timeout=15, verify=False, allow_redirects=True)
                response.raise_for_status()
        else:
             try:
                 response = session.get(url, timeout=15, verify=True, allow_redirects=True)
                 response.raise_for_status()
             except requests.exceptions.SSLError as ssl_err:
                 if verbose: print(f"[*] [{target_display}] SSL verification failed: {ssl_err}. Retrying without verification...")
                 response = session.get(url, timeout=15, verify=False, allow_redirects=True)
                 response.raise_for_status()
             except requests.exceptions.RequestException as req_err:
                 if verbose: print(f"[!] [{target_display}] Error fetching: {req_err}")
                 status = None
                 if hasattr(req_err, 'response') and req_err.response is not None: status = req_err.response.status_code
                 return None, None, url, None, status

        content = None
        try:
            response.encoding = response.apparent_encoding
            content = response.text
        except Exception as enc_err:
             if verbose: print(f"[!] [{target_display}] Warning: Encoding error: {enc_err}. Using response.text directly.")
             try: content = response.text
             except Exception as text_err:
                 if verbose: print(f"[!] [{target_display}] Error accessing response.text: {text_err}")
                 content = ""

        return response.headers, content, response.url, response.cookies, response.status_code

    except requests.exceptions.Timeout:
        if verbose: print(f"[!] [{target_display}] Error: Request timed out.")
        return None, None, url, None, "Timeout"
    except requests.exceptions.TooManyRedirects:
        if verbose: print(f"[!] [{target_display}] Error: Exceeded maximum redirects.")
        return None, None, url, None, "Too Many Redirects"
    except requests.exceptions.RequestException as e:
        if verbose: print(f"[!] [{target_display}] Error fetching: {e}")
        status = "Request Error";
        if hasattr(e, 'response') and e.response is not None: status = e.response.status_code
        return None, None, url, None, status
    except Exception as e:
        if verbose: print(f"[!] [{target_display}] An unexpected error occurred: {e}")
        return None, None, url, None, "Unexpected Error"
    finally:
        session.close()


# --- Detection Logic ---

def detect_cms(headers, html_content, url, cookies, verbose=False): # Added verbose
    """Detects CMS based on patterns, attempting to extract versions."""
    detected_cms = {} # Use dict: { 'CMS Name': 'version' or None }
    if not html_content: return detected_cms
    target_display = url[:50] + "..." if len(url) > 50 else url # For logging

    try:
        soup = BeautifulSoup(html_content, 'lxml') # Use lxml parser
    except Exception as e:
        if verbose: print(f"[!] [{target_display}] Error parsing HTML for CMS detection: {e}")
        return detected_cms

    # Pre-extract potential version from query parameters (?ver=...)
    query_param_version = None
    ver_pattern = re.compile(r'\?ver=([\d\.]+)')
    scripts_and_links = soup.select('script[src], link[rel="stylesheet"][href]')
    for tag in scripts_and_links:
        src_or_href = tag.get('src') or tag.get('href')
        if src_or_href:
            match = ver_pattern.search(src_or_href)
            if match:
                query_param_version = match.group(1)
                if verbose: print(f"[*] [{target_display}] Found potential version '{query_param_version}' from query parameter in {src_or_href}")
                break # Found one, stop looking

    for cms_name, patterns in CMS_PATTERNS.items():
        cms_identified = False
        version_found = None

        for pattern_info in patterns:
            ptype = pattern_info['type']
            match_found_in_pattern = False
            current_pattern_version = None

            try:
                if ptype == 'meta':
                    meta_tags = soup.select(f'meta[name="{pattern_info["name"]}"]')
                    for tag in meta_tags:
                        content = tag.get('content', '')
                        if content:
                            match = re.search(pattern_info['content'], content, re.IGNORECASE)
                            if match:
                                match_found_in_pattern = True
                                if 'version_group' in pattern_info and len(match.groups()) >= pattern_info['version_group'] and match.group(pattern_info['version_group']):
                                    current_pattern_version = match.group(pattern_info['version_group'])
                                    if verbose: print(f"[*] [{target_display}] CMS Version Match (Meta): {cms_name} {current_pattern_version}")
                                break
                elif ptype == 'path':
                    path_pattern = pattern_info['pattern']
                    elements = soup.select(f'a[href*="{path_pattern}"], link[href*="{path_pattern}"], script[src*="{path_pattern}"], img[src*="{path_pattern}"]')
                    if elements:
                        match_found_in_pattern = True
                        if verbose: print(f"[*] [{target_display}] CMS Pattern Match (Path): {cms_name} via '{path_pattern}'")
                elif ptype == 'header':
                    if headers is not None:
                        header_value = headers.get(pattern_info['name'])
                        if header_value:
                             match = re.search(pattern_info['pattern'], header_value, re.IGNORECASE)
                             if match:
                                match_found_in_pattern = True
                                if 'version_group' in pattern_info and len(match.groups()) >= pattern_info['version_group'] and match.group(pattern_info['version_group']):
                                    current_pattern_version = match.group(pattern_info['version_group'])
                                    if verbose: print(f"[*] [{target_display}] CMS Version Match (Header): {cms_name} {current_pattern_version}")
                elif ptype == 'cookie':
                     if cookies:
                        for cookie in cookies:
                            is_match = False
                            if isinstance(pattern_info['name'], re.Pattern):
                                if pattern_info['name'].match(cookie.name): is_match = True
                            elif pattern_info['name'] == cookie.name: is_match = True
                            if is_match: match_found_in_pattern = True; break
                elif ptype == 'script':
                     scripts = soup.find_all('script', src=True)
                     script_pattern_re = re.compile(pattern_info['pattern'], re.IGNORECASE)
                     for script in scripts:
                         if script_pattern_re.search(script['src']):
                             match_found_in_pattern = True
                             if verbose: print(f"[*] [{target_display}] CMS Pattern Match (Script): {cms_name} via '{pattern_info['pattern']}' in {script['src']}")
                             break
                elif ptype == 'html':
                     if isinstance(html_content, str) and re.search(pattern_info['pattern'], html_content, re.IGNORECASE | re.DOTALL):
                         match_found_in_pattern = True
                elif ptype == 'link_script_ver':
                     if query_param_version:
                         # Apply this version only if CMS is already identified but lacks version
                         if cms_identified and version_found is None:
                             current_pattern_version = query_param_version
                             match_found_in_pattern = True # Mark as match to update version
                             if verbose: print(f"[*] [{target_display}] CMS Version Match (Query Param): {cms_name} {current_pattern_version}")

                # --- Update results based on current pattern match ---
                if match_found_in_pattern:
                    cms_identified = True # Mark that we found evidence for this CMS
                    # Prioritize version found by this specific pattern (meta/header first)
                    if current_pattern_version:
                        version_found = current_pattern_version
                        # If version found via meta/header, we can often stop checking other patterns
                        if ptype in ['meta', 'header']:
                             break # Break inner loop

            except Exception as e:
                if verbose: print(f"[!] [{target_display}] Error during CMS detection pattern ({cms_name} - {ptype}): {e}")
                continue # Try next pattern

        # --- Finalize detection for this CMS ---
        if cms_identified:
            # Use the best version found (prioritizing meta/header, then query params)
            final_version = version_found if version_found else (query_param_version if cms_name == 'WordPress' else None) # Only apply query param version to WP for now
            detected_cms[cms_name] = final_version # Store name and best-found version (can be None)

    # --- Format output ---
    formatted_cms = []
    for name, ver in detected_cms.items():
        formatted_cms.append(f"{name} ({ver})" if ver else name)
    return formatted_cms


def detect_backend(headers, cookies, url, verbose=False): # Added verbose
    """Detects backend technologies based on headers and cookies, extracting versions."""
    detected_backend = {} # Use dict: { 'Tech Name': 'version' or None }
    if headers is None: headers = {}
    if cookies is None: cookies = []
    target_display = url[:50] + "..." if len(url) > 50 else url # For logging

    for tech_name, patterns in BACKEND_PATTERNS.items():
        if tech_name in detected_backend and detected_backend[tech_name] is not None: continue

        for pattern_info in patterns:
            ptype = pattern_info['type']
            version = None
            match_found = False
            try:
                if ptype == 'header':
                    header_value = headers.get(pattern_info['name'])
                    if header_value:
                        match = re.search(pattern_info['pattern'], header_value, re.IGNORECASE)
                        if match:
                            match_found = True
                            if 'version_group' in pattern_info and len(match.groups()) >= pattern_info['version_group'] and match.group(pattern_info['version_group']):
                                version = match.group(pattern_info['version_group'])
                elif ptype == 'cookie':
                    for cookie in cookies:
                        if pattern_info['name'] == cookie.name:
                            match_found = True; break

                if match_found:
                    if tech_name not in detected_backend or (version and detected_backend[tech_name] is None):
                         detected_backend[tech_name] = version
                    if version: break

            except Exception as e:
                if verbose: print(f"[!] [{target_display}] Error during Backend detection pattern ({tech_name} - {ptype}): {e}")
                continue

            if tech_name in detected_backend and detected_backend[tech_name]: break

    formatted_backend = []
    for name, ver in detected_backend.items():
        formatted_backend.append(f"{name} ({ver})" if ver else name)
    return formatted_backend


def detect_frontend(html_content, url, verbose=False): # Added verbose
    """Detects frontend libraries, attempting to extract versions from filenames/attributes."""
    detected_frontend = {} # Use dict: { 'Lib Name': 'version' or None }
    if not isinstance(html_content, str) or not html_content:
         return detected_frontend
    target_display = url[:50] + "..." if len(url) > 50 else url # For logging

    try:
        soup = BeautifulSoup(html_content, 'lxml') # Use lxml parser
    except Exception as e:
        if verbose: print(f"[!] [{target_display}] Error parsing HTML for Frontend detection: {e}")
        return detected_frontend

    scripts = soup.find_all('script')
    script_sources = [s.get('src') for s in scripts if s.get('src')]
    inline_scripts = ' '.join([s.string for s in scripts if s.string])
    css_links = [link.get('href') for link in soup.find_all('link', rel='stylesheet') if link.get('href')]

    for lib_name, patterns in FRONTEND_PATTERNS.items():
        if lib_name in detected_frontend and detected_frontend[lib_name] is not None: continue

        for pattern_info in patterns:
            ptype = pattern_info['type']
            version = None
            match_found = False
            try:
                if ptype == 'script':
                    for src in script_sources:
                        if isinstance(src, str):
                            match = re.search(pattern_info['pattern'], src, re.IGNORECASE)
                            if match:
                                match_found = True
                                if 'version_group' in pattern_info and len(match.groups()) >= pattern_info['version_group'] and match.group(pattern_info['version_group']):
                                    version = match.group(pattern_info['version_group'])
                                break
                elif ptype == 'html':
                     match = re.search(pattern_info['pattern'], html_content, re.IGNORECASE | re.DOTALL)
                     if match:
                        match_found = True
                        if 'version_group' in pattern_info and len(match.groups()) >= pattern_info['version_group'] and match.group(pattern_info['version_group']):
                            version = match.group(pattern_info['version_group'])
                elif ptype == 'script_content':
                    if inline_scripts:
                        match = re.search(pattern_info['pattern'], inline_scripts, re.IGNORECASE | re.DOTALL)
                        if match:
                            match_found = True
                            if 'version_group' in pattern_info and len(match.groups()) >= pattern_info['version_group'] and match.group(pattern_info['version_group']):
                                version = match.group(pattern_info['version_group'])
                elif ptype == 'css':
                     for href in css_links:
                         if isinstance(href, str):
                            match = re.search(pattern_info['pattern'], href, re.IGNORECASE)
                            if match:
                                match_found = True
                                if 'version_group' in pattern_info and len(match.groups()) >= pattern_info['version_group'] and match.group(pattern_info['version_group']):
                                    version = match.group(pattern_info['version_group'])
                                break

                if match_found:
                    if lib_name not in detected_frontend or (version and detected_frontend[lib_name] is None):
                         detected_frontend[lib_name] = version
                    if version: break

            except Exception as e:
                if verbose: print(f"[!] [{target_display}] Error during Frontend detection pattern ({lib_name} - {ptype}): {e}")
                continue

            if lib_name in detected_frontend and detected_frontend[lib_name]: break

    formatted_frontend = []
    for name, ver in detected_frontend.items():
        formatted_frontend.append(f"{name} ({ver})" if ver else name)
    return formatted_frontend


# --- WAF Detection using wafw00f ---
# (run_wafw00f remains unchanged)
def run_wafw00f(target_url, verbose):
    """Runs the wafw00f tool against the target URL and parses the output."""
    target_display = target_url[:50] + "..." if len(target_url) > 50 else target_url
    if verbose: print(f"[*] [{target_display}] Running wafw00f...")
    command = ['wafw00f', '-a', target_url]
    detected_waf = []
    generic_waf_detected = False # Flag for generic detection message

    # Define regex pattern to remove ANSI escape codes
    ansi_escape_pattern = re.compile(r'\x1b(?:[@-Z\\-_]|\[[0-?]*[ -/]*[@-~])')

    try:
        process = subprocess.run(command, capture_output=True, text=True, timeout=60, check=False, encoding='utf-8', errors='ignore')

        # Clean ANSI codes from stdout and stderr *before* parsing
        cleaned_stdout = ansi_escape_pattern.sub('', process.stdout)
        cleaned_stderr = ansi_escape_pattern.sub('', process.stderr)

        if process.returncode != 0:
            # Show "command not found" error regardless of verbose flag if -w was used
            if "command not found" in cleaned_stderr.lower() or "no such file" in cleaned_stderr.lower():
                 print(f"[!] [{target_display}] Error: 'wafw00f' command not found. Please ensure it's installed and in your PATH.")
                 return ["Error: wafw00f not found"] # Return specific error message
            # Show other errors only if verbose
            elif verbose:
                if "timed out" in cleaned_stderr.lower():
                     print(f"[!] [{target_display}] Error: wafw00f scan timed out.")
                else:
                     print(f"[!] [{target_display}] Error running wafw00f (return code {process.returncode}): {cleaned_stderr.strip()}")
            return [] # Return empty list for other errors unless verbose

        # Parse cleaned stdout for WAF information
        output_lines = cleaned_stdout.splitlines()
        specific_waf_pattern = re.compile(r'is behind\s+(?:the\s+)?([\w\s-]+?)(?:\s+\(.*\))?\s+WAF', re.IGNORECASE)
        generic_waf_pattern = re.compile(r'behind a WAF or some sort of security solution', re.IGNORECASE)
        found_waf_pattern = re.compile(r'\[\+\]\s+Found WAF:\s*([\w\s-]+?)(?:\s+\(.*\))?$', re.IGNORECASE)

        potential_waf_names = []
        for line in output_lines:
            found_match = found_waf_pattern.search(line)
            if found_match:
                waf_name = found_match.group(1).strip()
                potential_waf_names.append(waf_name)
                continue

            specific_match = specific_waf_pattern.search(line)
            if specific_match:
                waf_name = specific_match.group(1).strip()
                if "generic detection" not in waf_name.lower():
                     potential_waf_names.append(waf_name)

            elif generic_waf_pattern.search(line):
                 generic_waf_detected = True

        # Process potential names: handle "and/or", unique the list
        final_waf_list = set()
        for name in potential_waf_names:
            parts = re.split(r'\s+and(?:/or)?\s+|,', name, flags=re.IGNORECASE)
            for part in parts:
                cleaned_part = part.strip()
                cleaned_part = re.sub(r',\s*a$', '', cleaned_part).strip() # Remove trailing ', a'
                if cleaned_part:
                    final_waf_list.add(cleaned_part)

        detected_waf = list(final_waf_list)

        if not detected_waf and generic_waf_detected:
            detected_waf.append("Generic WAF/Security Solution")

        if not detected_waf and verbose:
             print(f"[*] [{target_display}] wafw00f: No WAF detected.")
        elif detected_waf and verbose:
             print(f"[*] [{target_display}] wafw00f detected: {', '.join(detected_waf)}")


    except FileNotFoundError:
        # Show "command not found" error regardless of verbose flag if -w was used
        print(f"[!] [{target_display}] Error: 'wafw00f' command not found. Please ensure it's installed and in your PATH.")
        return ["Error: wafw00f not found"] # Return specific error message
    except subprocess.TimeoutExpired:
        if verbose: print(f"[!] [{target_display}] Error: wafw00f scan timed out.")
        return []
    except Exception as e:
        if verbose: print(f"[!] [{target_display}] An unexpected error occurred while running wafw00f: {e}")
        return []

    return detected_waf

# --- Shodan Integration ---
def query_shodan(ip_address, api_key, verbose):
    """Queries Shodan API for host information."""
    global shodan_api_error_printed # Use global flag
    if not ip_address:
        return None # Cannot query without IP

    try:
        import shodan
    except ImportError:
        # Print only once if library missing
        if not shodan_api_error_printed:
            print("[!] Error: 'shodan' library not found. Please install it (`pip install shodan`) to use Shodan integration.")
            shodan_api_error_printed = True
        return None

    if verbose: print(f"[*] Querying Shodan for IP: {ip_address}")

    try:
        api = shodan.Shodan(api_key)
        host_info = api.host(ip_address)
        # Extract relevant data
        shodan_data = {
            "IP Address": ip_address, # Add IP address to results
            "Organization": host_info.get('org', 'N/A'),
            "ISP": host_info.get('isp', 'N/A'),
            "Hostnames": ", ".join(host_info.get('hostnames', [])),
            "Open Ports": ", ".join(map(str, host_info.get('ports', []))),
            "Tags": ", ".join(host_info.get('tags', [])), # Often contains technologies
            "Vulnerabilities": ", ".join(host_info.get('vulns', [])) # CVEs if found
        }
        if verbose: print(f"[*] Shodan data found for {ip_address}")
        return shodan_data
    except shodan.APIError as e:
        # Print specific API errors (like invalid key) only once
        error_msg = str(e).lower()
        if 'invalid api key' in error_msg or 'access restricted' in error_msg:
            if not shodan_api_error_printed:
                print(f"[!] Shodan API Error: {e}. Check your key in config.json.")
                print("[!] Note: Shodan free tier API keys have very limited access.")
                print("[!] Shodan queries will be skipped for the rest of this run.")
                shodan_api_error_printed = True
        elif verbose: # Print other API errors only if verbose
            print(f"[!] Shodan API error for {ip_address}: {e}")
        return None # Return None on API error
    except Exception as e:
        # Handle other potential errors (network, etc.)
        if verbose: print(f"[!] Error querying Shodan for {ip_address}: {e}")
        return None # Return None on other errors


# --- Config File Handling ---
def load_shodan_key(config_file="config.json"):
    """Loads Shodan API key from config file."""
    try:
        with open(config_file, 'r') as f:
            config_data = json.load(f)
            key = config_data.get("shodan_api_key")
            if key:
                print("[*] Loaded Shodan API key from config.json")
                return key
            else:
                # Inform user if file exists but key is missing
                print(f"[*] '{config_file}' found, but 'shodan_api_key' is missing or empty. Shodan queries disabled.")
                return None
    except FileNotFoundError:
        # It's okay if the file doesn't exist
        return None
    except json.JSONDecodeError:
        print(f"[!] Error: Could not decode JSON from {config_file}. Shodan queries disabled.")
        return None
    except Exception as e:
        print(f"[!] Error reading {config_file}: {e}. Shodan queries disabled.")
        return None


# --- Output Functions ---

def display_terminal(results):
    """Displays results for a single target in a consolidated format."""
    target_url = results.get('Target URL', 'N/A')
    final_url = results.get('Final URL', 'N/A')
    status_code = results.get('Status Code', 'N/A')

    cms_results = results.get('CMS', [])
    backend_results = results.get('Backend', [])
    frontend_results = results.get('Frontend', [])
    waf_results = results.get('WAF', [])
    shodan_results = results.get('Shodan', None) # Get Shodan results

    print(f"\n--- Results for: {target_url} ---")
    print(f"{'Final URL':<12}: {final_url} (Status: {status_code})")

    cms_str = ', '.join(cms_results) if cms_results else NOT_DETECTED_STRING # Use constant
    backend_str = ', '.join(backend_results) if backend_results else NOT_DETECTED_STRING # Use constant
    frontend_str = ', '.join(frontend_results) if frontend_results else NOT_DETECTED_STRING # Use constant
    if results.get('WAF_Checked') == False: waf_str = "Not Checked"
    elif waf_results: waf_str = ', '.join(waf_results)
    else: waf_str = "No Waf"

    print(f"{'CMS':<12}: {cms_str}")
    print(f"{'Backend':<12}: {backend_str}")
    print(f"{'Frontend':<12}: {frontend_str}")
    print(f"{'WAF':<12}: {waf_str}")

    # Display Shodan Info only if it's a dictionary (successful query)
    if isinstance(shodan_results, dict): # Check it's dict (implicitly excludes None and error strings)
        print("\n[+] Shodan Info:")
        for key, value in shodan_results.items():
            display_value = value
            if isinstance(value, str) and len(value) > 80: display_value = value[:77] + "..."
            if display_value: print(f"  {key:<15}: {display_value}")
    elif results.get('Shodan_Checked'): # If checked but failed/no data
        print(f"{'Shodan':<12}: Query attempted, no data or error.")


    print("-" * (len(target_url) + 18)) # Separator


def save_excel(all_results, filename):
    """Saves the results for all targets to an Excel file, merging cells and adding borders."""
    if not all_results:
        print("[!] No results to save.")
        return

    try:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])

        # Define styles
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal='center', vertical='center')
        left_vertical_center_alignment = Alignment(horizontal='left', vertical='center')
        thin_side = Side(style='thin')
        medium_side = Side(style='medium')
        summary_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
        domain_separator_border = Border(top=medium_side)

        # --- Calculate Status Code Summary ---
        status_codes = [str(result.get('Status Code', 'N/A')) for result in all_results]
        status_counts = Counter(status_codes)
        summary_rows_needed = 1 + len(status_counts) + 1
        summary_last_row = summary_rows_needed

        # --- Scan Summary Sheet ---
        ws_summary = wb.create_sheet("Scan Summary")
        ws_summary.insert_rows(1, amount=summary_rows_needed)

        # Write Status Code Summary Title
        summary_title_cell = ws_summary.cell(row=1, column=1, value="Status Code Summary")
        summary_title_cell.font = bold_font
        summary_title_cell.alignment = center_alignment
        ws_summary.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

        # Write Status Code Counts
        current_summary_row = 2
        sorted_status_codes = sorted(status_counts.keys(), key=lambda x: (int(x) if x.isdigit() else float('inf'), x))
        for code in sorted_status_codes:
            count = status_counts[code]
            ws_summary.cell(row=current_summary_row, column=1, value=f"{code}:")
            ws_summary.cell(row=current_summary_row, column=2, value=count)
            current_summary_row += 1

        # Apply Border around the summary section
        summary_data_last_row = summary_last_row - 1
        for row_idx in range(1, summary_data_last_row + 1):
            for col_idx in range(1, 3):
                cell = ws_summary.cell(row=row_idx, column=col_idx)
                cell.border = summary_border

        # Write Main Summary Table Header
        main_header_row = summary_last_row + 1
        ws_summary.cell(row=main_header_row, column=1, value="Target URL").font = bold_font
        ws_summary.cell(row=main_header_row, column=2, value="Final URL").font = bold_font
        ws_summary.cell(row=main_header_row, column=3, value="Status Code").font = bold_font

        # Write Main Summary Data
        current_data_row = main_header_row + 1
        for result in all_results:
             ws_summary.cell(row=current_data_row, column=1, value=result.get('Target URL', 'N/A'))
             ws_summary.cell(row=current_data_row, column=2, value=result.get('Final URL', 'N/A'))
             ws_summary.cell(row=current_data_row, column=3, value=result.get('Status Code', 'N/A'))
             current_data_row += 1

        # Auto-adjust column widths for summary sheet
        for col_idx in range(1, ws_summary.max_column + 1):
             max_length = 0; column = get_column_letter(col_idx)
             for row_idx in range(1, ws_summary.max_row + 1):
                 cell = ws_summary.cell(row=row_idx, column=col_idx)
                 try:
                     is_merged_title = False
                     if row_idx == 1 and col_idx == 1 and ws_summary.cell(row=1, column=1).coordinate in ws_summary.merged_cells: is_merged_title = True
                     val_str = str(cell.value) if cell.value is not None else ""
                     if is_merged_title: val_str = str(ws_summary.cell(row=1, column=1).value)
                     if len(val_str) > max_length: max_length = len(val_str)
                 except: pass
             ws_summary.column_dimensions[column].width = (max_length + 2)


        # --- Consolidated Technologies Sheet ---
        ws_tech = wb.create_sheet("Technologies")
        # Add WAF column to header -> New Header Order
        ws_tech.append(["Target URL", "WAF", "Category", "Technology (Version)"])
        for cell in ws_tech["1:1"]: cell.font = bold_font

        current_row = 2
        url_merge_start_row = 2

        for idx, result in enumerate(all_results):
            target = result.get('Target URL', 'N/A')
            row_start_for_this_target = current_row

            # Determine WAF string for this target *once*
            waf_list = result.get('WAF', [])
            if result.get('WAF_Checked') == False:
                waf_display_str = "Not Checked"
            elif waf_list:
                waf_display_str = ', '.join(waf_list)
            else:
                waf_display_str = "No Waf"

            # Apply Top Border if New Domain
            if idx > 0:
                prev_target = all_results[idx-1].get('Target URL', 'N/A')
                if target != prev_target:
                    for col_idx in range(1, 5): # Columns A, B, C, D
                         cell = ws_tech.cell(row=row_start_for_this_target, column=col_idx)
                         existing_border = cell.border if cell.border else Border()
                         cell.border = Border(top=domain_separator_border.top, left=existing_border.left, right=existing_border.right, bottom=existing_border.bottom)

            # --- Write Data and Track Merges ---
            rows_added_for_target = 0 # Track rows added for this target for merging

            # Define categories including Shodan (only if Shodan data exists and is not an error)
            shodan_data = result.get('Shodan')
            # Check if shodan_data is a dictionary and does not contain an 'Error' key
            shodan_items_to_write = []
            add_shodan_category = False
            if isinstance(shodan_data, dict) and "Error" not in shodan_data: # Check it's a dict AND no error key
                shodan_items_to_write = [f"{k}: {v}" for k, v in shodan_data.items() if v and v != 'N/A']
                if shodan_items_to_write: # Only add category if there's actual data
                    add_shodan_category = True

            categories_data = {
                "CMS": result.get('CMS', []),
                "Backend": result.get('Backend', []),
                "Frontend": result.get('Frontend', []),
                "Shodan": shodan_items_to_write # Use the processed list
            }

            # Determine which categories to include in the loop
            categories_to_process = ["CMS", "Backend", "Frontend"]
            if add_shodan_category: # Only add Shodan if data is valid
                categories_to_process.append("Shodan")

            # Add rows for each category
            for category in categories_to_process:
                items = categories_data.get(category, [])
                if items: # If list is not empty
                    category_merge_start_row = current_row
                    for item in items:
                        ws_tech.cell(row=current_row, column=1, value=target)
                        ws_tech.cell(row=current_row, column=2, value=waf_display_str)
                        ws_tech.cell(row=current_row, column=3, value=category)
                        ws_tech.cell(row=current_row, column=4, value=item)
                        current_row += 1
                        rows_added_for_target += 1
                    # Merge Category cells if needed
                    category_merge_end_row = current_row - 1
                    if category_merge_end_row > category_merge_start_row:
                        try:
                            ws_tech.merge_cells(start_row=category_merge_start_row, start_column=3, end_row=category_merge_end_row, end_column=3)
                            merged_cell = ws_tech.cell(row=category_merge_start_row, column=3)
                            merged_cell.alignment = left_vertical_center_alignment
                        except Exception as merge_err: print(f"[!] Warning: Could not merge Category cells for '{target}' > '{category}'. Error: {merge_err}")
                else: # If list is empty add "Not Found" row
                    ws_tech.cell(row=current_row, column=1, value=target)
                    ws_tech.cell(row=current_row, column=2, value=waf_display_str)
                    ws_tech.cell(row=current_row, column=3, value=category)
                    ws_tech.cell(row=current_row, column=4, value=NOT_DETECTED_STRING) # Use constant
                    current_row += 1
                    rows_added_for_target += 1


            # If no rows were added at all (e.g., error + no WAF check + no Shodan)
            if rows_added_for_target == 0:
                 ws_tech.cell(row=current_row, column=1, value=target)
                 ws_tech.cell(row=current_row, column=2, value=waf_display_str)
                 ws_tech.cell(row=current_row, column=3, value="N/A")
                 ws_tech.cell(row=current_row, column=4, value="No technologies detected")
                 current_row += 1


            # --- URL & WAF Cell Merging Logic ---
            is_last_result = (idx == len(all_results) - 1)
            next_target_is_different = False
            if not is_last_result:
                next_target = all_results[idx + 1].get('Target URL', 'N/A')
                if target != next_target: next_target_is_different = True

            if next_target_is_different or is_last_result:
                url_merge_end_row = current_row - 1
                if url_merge_end_row >= url_merge_start_row:
                    # Check if more than one row was *actually added* for this target before merging
                    if rows_added_for_target > 1: # Check if we added more than one row
                        try:
                            ws_tech.merge_cells(start_row=url_merge_start_row, start_column=1, end_row=url_merge_end_row, end_column=1)
                            merged_cell_url = ws_tech.cell(row=url_merge_start_row, column=1)
                            merged_cell_url.alignment = left_vertical_center_alignment
                            ws_tech.merge_cells(start_row=url_merge_start_row, start_column=2, end_row=url_merge_end_row, end_column=2)
                            merged_cell_waf = ws_tech.cell(row=url_merge_start_row, column=2)
                            merged_cell_waf.alignment = left_vertical_center_alignment
                        except Exception as merge_err:
                            print(f"[!] Warning: Could not merge URL/WAF cells for target '{target}' (rows {url_merge_start_row}-{url_merge_end_row}). Error: {merge_err}")
                    url_merge_start_row = current_row


        # Auto-adjust column widths for technologies sheet
        for col in ws_tech.columns:
             max_length = 0; column = get_column_letter(col[0].column)
             for cell in col:
                 try:
                     is_merged = False; is_top_left = False
                     for merged_range in ws_tech.merged_cells.ranges:
                         if (merged_range.min_row <= cell.row <= merged_range.max_row and
                             merged_range.min_col <= cell.column <= merged_range.max_col):
                             is_merged = True
                             if cell.row == merged_range.min_row and cell.column == merged_range.min_col: is_top_left = True
                             break
                     val_str = str(cell.value) if cell.value is not None else ""
                     if not is_merged or is_top_left:
                          if len(val_str) > max_length: max_length = len(val_str)
                 except: pass
             adjusted_width = (max_length + 2); min_width = 15
             if column in ['B', 'D']: min_width = 25
             elif column == 'C': min_width = 10
             ws_tech.column_dimensions[column].width = max(adjusted_width, min_width)


        wb.save(filename)
        print(f"\n[+] Results for {len(all_results)} target(s) successfully saved to {filename}")

    except NameError as e:
        if 'openpyxl' in str(e): print("[!] Error: `openpyxl` library not found or not imported correctly.\n[-] Please ensure it's installed: pip install openpyxl")
        elif 'shodan' in str(e): print("[!] Error: `shodan` library not found or not imported correctly.\n[-] Please ensure it's installed: pip install shodan")
        else: print(f"[!] An unexpected NameError occurred: {e}")
    except Exception as e:
        print(f"[!] Error saving Excel file {filename}: {e}")


# --- Threading Worker ---
def scan_target_worker(q, results_list, results_lock, check_waf, shodan_key, verbose): # Added shodan_key
    """Worker thread function to process URLs from the queue."""
    target_url = None
    try:
        while not q.empty():
            try:
                target_url = q.get_nowait()
            except queue.Empty:
                continue

            target_display = target_url[:50] + "..." if len(target_url) > 50 else target_url
            if verbose: print(f"[*] Thread {threading.current_thread().name} scanning: {target_display}")

            headers, html_content, final_url, cookies, status_code = make_request(target_url, verbose) # Pass verbose

            current_result = {
                'Target URL': target_url,
                'Final URL': final_url if final_url else target_url,
                'Status Code': status_code,
                'CMS': [], 'Backend': [], 'Frontend': [], 'WAF': [], 'Shodan': None, # Initialize Shodan
                'WAF_Checked': check_waf,
                'Shodan_Checked': bool(shodan_key) # Track if Shodan check was requested and key was valid
            }

            is_error_status = status_code is None or not isinstance(status_code, int) or status_code >= 400
            no_content = not headers and not html_content

            # --- Run Optional Checks ---
            if check_waf:
                current_result['WAF'] = run_wafw00f(target_url, verbose)

            shodan_ip_address = None # Track IP used for Shodan
            if shodan_key: # Only query if key is provided
                try:
                    hostname_to_resolve = urlparse(final_url if final_url else target_url).hostname
                    if hostname_to_resolve:
                        shodan_ip_address = socket.gethostbyname(hostname_to_resolve) # Store the IP
                        current_result['Shodan'] = query_shodan(shodan_ip_address, shodan_key, verbose)
                    elif verbose: print(f"[!] [{target_display}] Could not extract hostname for Shodan lookup.")
                except socket.gaierror as e:
                    if verbose: print(f"[!] [{target_display}] DNS resolution failed for Shodan lookup: {e}")
                except Exception as e:
                     if verbose: print(f"[!] [{target_display}] Error during DNS resolution for Shodan: {e}")
                # Shodan result will be None if query fails


            # --- Perform Core Detections ---
            if is_error_status:
                if not no_content:
                     current_result['CMS'] = detect_cms(headers, html_content, target_url, cookies, verbose)
                     current_result['Backend'] = detect_backend(headers, cookies, target_url, verbose)
                     current_result['Frontend'] = detect_frontend(html_content, target_url, verbose)
            else:
                 current_result['CMS'] = detect_cms(headers, html_content, target_url, cookies, verbose)
                 current_result['Backend'] = detect_backend(headers, cookies, target_url, verbose)
                 current_result['Frontend'] = detect_frontend(html_content, target_url, verbose)

            # Safely append result
            with results_lock:
                results_list.append(current_result)

            q.task_done()
    except Exception as e:
         target_display = target_url[:50] + "..." if target_url and len(target_url) > 50 else target_url
         print(f"[!] Thread {threading.current_thread().name} encountered an error processing '{target_display}': {e}")
         if target_url:
              try: q.task_done()
              except ValueError: pass


# --- Main Execution ---
def main():
    parser = argparse.ArgumentParser(
        description="BackRecon Tool: Detect web technologies (CMS, Backend, Frontend), optionally WAF and Shodan info.", # Updated desc
        epilog="Notes:\n"
               "  WAF detection (-w) requires the 'wafw00f' tool to be installed and in your system's PATH.\n"
               "  Shodan integration (-s) requires a 'config.json' file with your API key and the 'shodan' Python library (`pip install shodan`).\n" # Updated Shodan note
               "  Example config.json: {\"shodan_api_key\": \"YOUR_KEY_HERE\"}"
    )
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("-d", "--domain", help="Single target domain or URL (e.g., example.com)")
    group.add_argument("-l", "--list", help="File containing a list of target domains or URLs (one per line).")
    parser.add_argument("-o", "--output", help="Output filename for combined Excel report (e.g., report.xlsx).")
    parser.add_argument("-t", "--threads", type=int, default=4, help="Number of concurrent threads for list scanning (default: 4).")
    parser.add_argument("-w", "--waf", action="store_true", help="Enable WAF detection using 'wafw00f'.")
    # Added -s flag for Shodan
    parser.add_argument("-s", "--shodan", action="store_true", help="Enable Shodan host enrichment (requires config.json with API key and 'shodan' library).")
    parser.add_argument("-v", "--verbose", action="store_true", help="Increase output verbosity (show errors, thread activity, etc.).")

    args = parser.parse_args()

    if args.threads < 1:
        print("[!] Error: Number of threads must be at least 1."); exit(1)

    output_filename = args.output
    if output_filename and not output_filename.lower().endswith('.xlsx'):
        output_filename += '.xlsx'
        print(f"[*] Output filename amended to: {output_filename}")

    # --- Load Shodan Key ---
    shodan_api_key = None
    if args.shodan: # Only try loading if -s flag is present
        shodan_api_key = load_shodan_key()
        if not shodan_api_key:
            print("[!] Shodan lookup enabled (-s) but no valid API key found in config.json. Shodan queries will be skipped.")
            # Proceed without Shodan key if not found/invalid

    # --- Dependency Check ---
    missing_libs = []
    try: import requests
    except ImportError: missing_libs.append("requests")
    try: import bs4
    except ImportError: missing_libs.append("beautifulsoup4")
    try: import openpyxl
    except ImportError: missing_libs.append("openpyxl")
    try: import lxml
    except ImportError: missing_libs.append("lxml")
    # Check for shodan only if -s flag is used and key was potentially loaded
    if args.shodan:
        try: import shodan
        except ImportError: missing_libs.append("shodan")


    if missing_libs:
        print(f"[!] Missing required libraries: {', '.join(missing_libs)}")
        print(f"[-] Please install them using: pip install {' '.join(missing_libs)}")
        exit(1)
    # --- End Dependency Check ---

    targets = []
    if args.domain:
        targets.append(args.domain.strip())
    elif args.list:
        if not os.path.exists(args.list):
             print(f"[!] Error: Input file not found: {args.list}"); exit(1)
        try:
            with open(args.list, 'r', encoding='utf-8') as f:
                targets = [line.strip() for line in f if line.strip() and not line.startswith('#')]
            if not targets:
                 print(f"[!] Error: Input file '{args.list}' is empty or contains no valid targets."); exit(1)
            print(f"[*] Loaded {len(targets)} target(s) from {args.list}")
        except IOError as e:
            print(f"[!] Error reading file {args.list}: {e}"); exit(1)
        except Exception as e:
             print(f"[!] Error processing file {args.list}: {e}"); exit(1)


    all_scan_results = []
    results_lock = threading.Lock()

    use_threading = args.list and args.threads > 1 and len(targets) > 1
    # Determine if Shodan should actually be checked (flag + valid key)
    check_shodan = args.shodan and shodan_api_key

    if use_threading:
        print(f"[*] Starting scan with {args.threads} threads...")
        target_queue = queue.Queue()
        for target in targets:
            target_queue.put(target)

        threads = []
        for i in range(args.threads):
            # Pass the *actual* key if check_shodan is true, else None
            shodan_key_to_pass = shodan_api_key if check_shodan else None
            thread = threading.Thread(target=scan_target_worker, args=(target_queue, all_scan_results, results_lock, args.waf, shodan_key_to_pass, args.verbose), name=f"Worker-{i+1}")
            thread.daemon = True
            threads.append(thread)
            thread.start()

        target_queue.join()
        print("[*] Threaded scan complete.")

    else:
        # Sequential scan
        print(f"[*] Starting sequential scan...")
        for target_url in targets:
            target_display = target_url[:50] + "..." if len(target_url) > 50 else target_url

            headers, html_content, final_url, cookies, status_code = make_request(target_url, args.verbose)

            current_result = {
                'Target URL': target_url,
                'Final URL': final_url if final_url else target_url,
                'Status Code': status_code,
                'CMS': [], 'Backend': [], 'Frontend': [], 'WAF': [], 'Shodan': None, # Init Shodan
                'WAF_Checked': args.waf,
                'Shodan_Checked': check_shodan # Track if Shodan check was requested and key was valid
            }

            is_error_status = status_code is None or not isinstance(status_code, int) or status_code >= 400
            no_content = not headers and not html_content

            # --- Run Optional Checks ---
            if args.waf:
                current_result['WAF'] = run_wafw00f(target_url, args.verbose)

            shodan_ip_address = None # Track IP used for Shodan
            if check_shodan: # Only run if flag is set AND key is valid
                try:
                    hostname_to_resolve = urlparse(final_url if final_url else target_url).hostname
                    if hostname_to_resolve:
                        shodan_ip_address = socket.gethostbyname(hostname_to_resolve)
                        current_result['Shodan'] = query_shodan(shodan_ip_address, shodan_api_key, args.verbose)
                    elif args.verbose: print(f"[!] [{target_display}] Could not extract hostname for Shodan lookup.")
                except socket.gaierror as e:
                    if args.verbose: print(f"[!] [{target_display}] DNS resolution failed for Shodan lookup: {e}")
                except Exception as e:
                     if args.verbose: print(f"[!] [{target_display}] Error during DNS resolution for Shodan: {e}")
                # No need to add IP if query fails, query_shodan returns None now


            # --- Perform Core Detections ---
            if is_error_status:
                if not no_content:
                     current_result['CMS'] = detect_cms(headers, html_content, target_url, cookies, args.verbose)
                     current_result['Backend'] = detect_backend(headers, cookies, target_url, args.verbose)
                     current_result['Frontend'] = detect_frontend(html_content, target_url, args.verbose)
            else:
                 current_result['CMS'] = detect_cms(headers, html_content, target_url, cookies, args.verbose)
                 current_result['Backend'] = detect_backend(headers, cookies, target_url, args.verbose)
                 current_result['Frontend'] = detect_frontend(html_content, target_url, args.verbose)

            all_scan_results.append(current_result)

            if not output_filename:
                display_terminal(current_result)

    # --- Output Results ---
    if use_threading:
        target_order = {url: i for i, url in enumerate(targets)}
        all_scan_results.sort(key=lambda res: target_order.get(res.get('Target URL'), float('inf')))


    if output_filename:
        save_excel(all_scan_results, output_filename)
    elif not use_threading:
        print(f"\n[*] Scan finished for {len(targets)} target(s).")
    else:
         print("\n--- Scan Complete ---")
         print("[*] Displaying results for all targets:")
         for result in all_scan_results:
              display_terminal(result)
         print(f"\n[*] Scan finished for {len(targets)} target(s).")


if __name__ == "__main__":
    main()
